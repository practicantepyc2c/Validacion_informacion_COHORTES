from operator import index
from os import remove
from tokenize import String
from turtle import pos
import pandas as pd
import numpy as np
import os
from pathlib import Path
pd.options.display.max_rows = None
pd.options.display.max_columns = None
import time
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
parametrica_identificacion = "./Parametricas/tipos_identificacion.xlsx"
parametrica_codigos = "./Parametricas/codigos.xlsx"
parametrica_programa = "./Parametricas/Programas_Cohortes.xlsx"
from correo import envioCorreoError, envioCorreoCorrecto,envioCorreo_archivonopermitido
from obtenerRemitente import ultimoCorreo
from escribirLog import auditorio_cohortes
from shutil import copyfile



def validacionInfo(fichero,nombreArchivo):
	
	# Obtener correo del que envio el archivo
	remitente = ultimoCorreo(nombreArchivo)
	print('Valindandose...')
	# Leer archivo excel con pandas
	try:
		archivo = pd.read_excel(fichero)
	except:
		envioCorreo_archivonopermitido(nombreArchivo,remitente)
	# Leer parametricas
	tipos_identificacion = pd.read_excel(parametrica_identificacion)
	programas_cohortes = pd.read_excel(parametrica_programa)
	codigosPermitidos = pd.read_excel(parametrica_codigos)
	archivo


	# validar tipo de codigo segun parametrica codigosPermitidos
	codigo = archivo['Cod_CIE10']
	Cod_CIE10Permitido = archivo['Cod_CIE10'].isin(codigosPermitidos.Codigos_permitidos)
	noPermitido = Cod_CIE10Permitido.loc[Cod_CIE10Permitido == False]
	columnas_incorrecto_Cod_CIE10 = np.array(noPermitido.index.values)
	# Se suma dos a cada posicion para que sea igual a la posicion del archivo (los Dataframes comienzan desde el index 0)
	columnas_incorrecto_Cod_CIE10 = columnas_incorrecto_Cod_CIE10 + 2
	# Se guardan los errores en un diccionario 
	diccionario_errores = {columna:" cod_CIE10 incorrecto " for columna in columnas_incorrecto_Cod_CIE10}



	# Validar tipo de documento segun parametrico tipos_identificacion
	tipodoc = archivo['Tipo_Doc']
	tipo_docPermitido = archivo['Tipo_Doc'].isin(tipos_identificacion.tipo_identificacion)
	noPermitido = tipo_docPermitido.loc[tipo_docPermitido == False]
	columnas_incorrecto_tipo_doc = np.array(noPermitido.index.values)
	columnas_incorrecto_tipo_doc = columnas_incorrecto_tipo_doc + 2
	for posicion in columnas_incorrecto_tipo_doc:
		# Valida si en la fila ya existe un error si es asi agrega el nuevo error encontrado
		if posicion in diccionario_errores:
			nuevoError = diccionario_errores[posicion] + '-' + ' Tipo_Doc incorrecto '
			diccionario_errores.update({posicion:nuevoError})
		else:
			diccionario_errores.update({posicion: ' Tipo_Doc incorrecto '})
	
	# Validar Num_doc, sin caracteres especiales y con menos de 17 digitos
	# Caracteres especiales
	caracteres_especiales = ['|','°','¬','!','"','#','$','%','&','/','(',')','=','?','¿','¡','@','´','¨','*','+','~','}',']','`','{','[','^',',',';',':','.','-','_','<','>']
	num_doc = archivo['Num_Doc']
	num_doc = num_doc.fillna('Nulo')
	columnas_incorrecto_num_doc = []
	for i in range(len(num_doc)):
		for j in range(len(caracteres_especiales)):
			# Transformacion en string para usar funcion find()
			valor_str = str(num_doc[i])
			# encontrar caracter especial en el valor actual y validar la cantidad de digitos
			if valor_str.find(caracteres_especiales[j]) != -1 or len(valor_str) > 17 or valor_str == 'Nulo':
				columnas_incorrecto_num_doc.append(i)
				break
	columnas_incorrecto_num_doc = np.array(columnas_incorrecto_num_doc)
	columnas_incorrecto_num_doc = columnas_incorrecto_num_doc + 2
	for posicion in columnas_incorrecto_num_doc:
		if posicion in diccionario_errores:
			nuevoError = diccionario_errores[posicion] + '-' + ' Num_Doc incorrecto '
			diccionario_errores.update({posicion:nuevoError})
		else:
			diccionario_errores.update({posicion: ' Num_Doc incorrecto '})


	# Validar programa segun parametrica programas_cohortes
	programa_permitido = archivo['Programa'].isin(programas_cohortes.PROGRAMAS)
	programa_nopermitido = programa_permitido.loc[programa_permitido == False]
	columnas_incorrecto_programa = np.array(programa_nopermitido.index.values)
	columnas_incorrecto_programa = columnas_incorrecto_programa + 2
	for posicion in columnas_incorrecto_programa:
		if posicion in diccionario_errores:
			nuevoError = diccionario_errores[posicion] + '-' + ' Programa incorrecto '
			diccionario_errores.update({posicion:nuevoError})
		else:
			diccionario_errores.update({posicion: ' Programa incorrecto '})


	# Validar periodo segun el periodo actual
	periodo = archivo['Periodo']
	columnas_incorrecto_periodo = []
	# extraer el periodo correcto del archivo
	nombreArchivo_segmentado = nombreArchivo.split('-')
	periodo_anterior = nombreArchivo_segmentado[1]
	periodo_anterior = periodo_anterior.strip()
	nombre_nuevo = nombreArchivo_segmentado[0] + '-' + periodo_anterior + '.xlsx'
	print(nombre_nuevo)
	periodo_anterior = int(periodo_anterior)
	# Concatena año y mes para crear periodo actual
	for i in range(len(periodo)):
		if periodo[i] != periodo_anterior:
			columnas_incorrecto_periodo.append(i)
	columnas_incorrecto_periodo = np.array(columnas_incorrecto_periodo)
	columnas_incorrecto_periodo = columnas_incorrecto_periodo + 2
	for posicion in columnas_incorrecto_periodo:
		if posicion in diccionario_errores:
			nuevoError = diccionario_errores[posicion] + '-' + ' Periodo incorrecto '
			diccionario_errores.update({posicion:nuevoError})
		else:
			diccionario_errores.update({posicion: ' Periodo incorrecto '})

	# validar Sub clasificacion que sus valores siempre sean mayusculas
	sub_clasificacion = archivo['Sub clasificacion']
	columnas_incorrecto_subClasificacion = []
	for i in range(len(sub_clasificacion)):
		try:
			# transforma el valor actual en mayusculas para compararlo 
			letra_mayuscula = sub_clasificacion[i].upper()
			if letra_mayuscula != sub_clasificacion[i]:
				columnas_incorrecto_subClasificacion.append(i)
		except:
			pass
	columnas_incorrecto_subClasificacion = np.array(columnas_incorrecto_subClasificacion)
	columnas_incorrecto_subClasificacion = columnas_incorrecto_subClasificacion + 2
	for posicion in columnas_incorrecto_subClasificacion:
		if posicion in diccionario_errores:
			nuevoError = diccionario_errores[posicion] + '-' + ' sub clasificacion incorrecto '
			diccionario_errores.update({posicion:nuevoError})
		else:
			diccionario_errores.update({posicion: ' sub clasificacion incorrecto '})


	# Validar campo controlado validar si sus valores son CONTROLADO, NO APLICA y NO CONTROLADO 
	controlado = archivo['controlado']
	columnas_incorrecto_controlado = []
	controlado = controlado.fillna('Nulo')
	for i in range(len(controlado)):
		if controlado[i] != 'CONTROLADO' and controlado[i] != 'NO APLICA' and controlado[i] != 'NO CONTROLADO':
			columnas_incorrecto_controlado.append(i)
	columnas_incorrecto_controlado = np.array(columnas_incorrecto_controlado)
	columnas_incorrecto_controlado = columnas_incorrecto_controlado + 2
	for posicion in columnas_incorrecto_controlado:
		if posicion in diccionario_errores:
			nuevoError = diccionario_errores[posicion] + '-' + ' controlado incorrecto '
			diccionario_errores.update({posicion:nuevoError})
		else:
			diccionario_errores.update({posicion: ' controlado incorrecto '})

	# Validacion seguimiento validar si sus valores son SI, NO y NO APLICA
	seguimiento = archivo['Seguimiento Programa']
	columnas_incorrecto_seguimiento = []
	seguimiento = seguimiento.fillna('Nulo')
	for i in range(len(seguimiento)):
		if seguimiento[i] != 'SI' and seguimiento[i] != 'NO' and seguimiento[i] != 'NO APLICA':
			columnas_incorrecto_seguimiento.append(i)
	columnas_incorrecto_seguimiento = np.array(columnas_incorrecto_seguimiento)
	columnas_incorrecto_seguimiento = columnas_incorrecto_seguimiento + 2	
	for posicion in columnas_incorrecto_seguimiento:
		if posicion in diccionario_errores:
			nuevoError = diccionario_errores[posicion] + '-' + ' Seguimiento Programa incorrecto '
			diccionario_errores.update({posicion:nuevoError})
		else:
			diccionario_errores.update({posicion: ' Seguimiento Programa incorrecto '})

	# Validar fecha > actual Fecha_Diagnostico
	# dia/mes/año formato correcto
	fechas = archivo['Fecha_Diagnostico']
	formato_incorrecto = []
	fechaActual = datetime.now().strftime('%d/%m/%Y')
	fechaActual = time.strptime(fechaActual, "%d/%m/%Y")
	fechas = fechas.fillna('Nulo')
	for i in range(len(fechas)):
		try:
			if fechas[i] > datetime.now():
				print(fechas[i],datetime.now())
				print(fechas[i] > datetime.now())
				formato_incorrecto.append(i)
		except:
			if fechas[i] != 'Nulo':
				try:
					fechaFormateada = time.strptime(fechas[i], "%d/%m/%Y")
					if fechaFormateada > fechaActual:
						formato_incorrecto.append(i)
				except:
					formato_incorrecto.append(i)
	formato_incorrecto = np.array(formato_incorrecto)
	formato_incorrecto = formato_incorrecto + 2
	for posicion in formato_incorrecto:
		if posicion in diccionario_errores:
			nuevoError = diccionario_errores[posicion] + '-' + 'Fecha_diagnostico incorrecto'
			diccionario_errores.update({posicion:nuevoError})
		else:
			diccionario_errores.update({posicion: 'Fecha_diagnostico incorrecto'})


	print("Validacion finalizada")
	# Enviar correo dependiendo si el archivo es correcto
	if(len(columnas_incorrecto_Cod_CIE10) == 0 and len(columnas_incorrecto_controlado) == 0 and len(columnas_incorrecto_seguimiento) == 0 and len(columnas_incorrecto_tipo_doc) == 0 and len(formato_incorrecto) == 0 and len(columnas_incorrecto_num_doc) == 0) and len(columnas_incorrecto_periodo) == 0:
		destino = Path('Archivo_correctos',nombreArchivo)
		copyfile(fichero,destino)
		envioCorreoCorrecto(nombreArchivo,remitente)
		auditorio_cohortes(datetime.now().strftime('%d/%m/%Y'),remitente,"Carga exitosa",nombreArchivo)
	else:
		print("Escribiendo...")
		wb = load_workbook(fichero)
		page = wb.active
		page.cell(1,10,value='Errores')
		# Escribir errores del diccionario_errores al excel
		for error in diccionario_errores:
			page.cell(error,10,value=diccionario_errores.get(error))
		# Proceso de envio
		wb.save(filename=fichero)
		destino = Path('C:/Users/Camilo/OneDrive/Documents/TrabajoKeralty/Code/Validacion_informacion_COHORTES',nombre_nuevo)
		copyfile(fichero,destino)
		print("Terminado")
		envioCorreoError(nombreArchivo,remitente,nombre_nuevo
		)
		auditorio_cohortes(datetime.now().strftime('%d/%m/%Y'),remitente,"Carga con errores",nombreArchivo)
		os.remove(nombre_nuevo)
	return print("archivo validado exitosamente")

#validacionInfo("Parametricas/programa_FALLA CARDIACA_202207 - Wilmar Yidid Fracica Velasquez.xlsx","programa_FALLA CARDIACA_202207 - Wilmar Yidid Fracica Velasquez")
